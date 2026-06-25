"""钢筋台账统一蓝图：进场/调拨/措施筋/盘点/废料/形象进度"""
import os
from datetime import datetime
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify
from flask_login import login_required, current_user
from sqlalchemy import extract, func
from app import db
from app.models import Project, Incoming, Transfer, MeasureRebar, Inventory, Waste, ImportedFile
from app.models.bom import Building as Bldg, Floor, Area, Component, RebarDetail
from app.services.pagination_service import paginate, get_page

ledger_bp = Blueprint("ledger", __name__, template_folder="../templates/ledger")


def _bc(*items):
    """构造统一面包屑：首页 + 各级节点"""
    base = [{"name": "钢筋管理平台", "url": url_for("home.index")}]
    base.extend(items)
    return base


UNIT_LABELS = {"kg": "公斤(kg)", "T": "吨(T)"}

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads", "ledger")


@ledger_bp.route("/<int:project_id>/import/<dtype>", methods=["GET", "POST"])
@login_required
def import_data(project_id, dtype, active_menu="file_management", active_submenu=None):
    """通用导入入口：incoming/transfer/measure/inventory/waste + 扩展类型"""
    from app.services.ledger_import import (
        IMPORTERS, save_upload, ALLOWED_EXT as ALLOWED,
        process_excel_import, DTYPE_TO_LEDGER_TYPE,
    )
    from app.routes.project import SHEET_LEDGER_MAP

    # 支持的台账类型（含占位类型）
    SUPPORTED_DTYPES = set(IMPORTERS.keys()) | {'detailing', 'non_budget', 'pile_foundation', 'support_structure'}
    REVERSE_SHEET_MAP = {v: k for k, v in SHEET_LEDGER_MAP.items()}

    p = get_project(project_id)
    if dtype not in SUPPORTED_DTYPES:
        flash("不支持的数据类型。", "danger")
        return redirect(url_for("project.list"))

    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            msg = "请选择文件。"
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"error": msg}), 400
            flash(msg, "warning")
            return redirect(url_for("ledger.import_data", project_id=project_id, dtype=dtype))

        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED:
            msg = "仅支持 .xlsx/.xls 文件。"
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"error": msg}), 400
            flash(msg, "danger")
            return redirect(url_for("ledger.import_data", project_id=project_id, dtype=dtype))

        try:
            # 保存文件
            path = save_upload(file, os.path.join(UPLOAD_DIR, dtype))

            # 创建 ImportedFile 记录（使用标准 ledger_type）
            ledger_type = DTYPE_TO_LEDGER_TYPE.get(dtype, dtype)
            imported = ImportedFile(
                project_id=project_id,
                ledger_type=ledger_type,
                original_filename=file.filename,
                file_path=path,
                file_size=os.path.getsize(path),
                uploaded_by=current_user.id,
                status="pending",
                progress=0,
            )
            db.session.add(imported)
            db.session.commit()

            # 异步投递 Celery 任务
            task = process_excel_import.delay(project_id, path, dtype, imported.id, current_user.id)

            # 回写 task_id
            imported.task_id = task.id
            db.session.commit()

            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({
                    "task_id": task.id,
                    "imported_file_id": imported.id,
                    "filename": file.filename,
                    "status": "pending",
                })

            flash(f"文件《{file.filename}》已提交，正在后台处理...", "info")
        except Exception as e:
            msg = f"提交失败：{e}"
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"error": msg}), 500
            flash(msg, "danger")

        # 非 AJAX 回退：重定向到列表页或 sheet 页
        if dtype in IMPORTERS:
            return redirect(url_for(f"ledger.{dtype}_list", project_id=project_id))
        else:
            sheet_view = REVERSE_SHEET_MAP.get(dtype)
            if sheet_view:
                return redirect(url_for(f"project.{sheet_view}"))
            return redirect(url_for("project.dashboard"))

    labels = {
        'incoming': '进场台账', 'transfer': '调拨台账',
        'measure': '措施筋台账', 'inventory': '盘点表', 'waste': '废料台账',
        'detailing': '钢筋翻样表', 'non_budget': '非预算收入使用钢筋表',
        'pile_foundation': '主体桩基表', 'support_structure': '基坑支护表',
    }

    # 计算返回列表的 URL（5 种核心类型走 ledger 列表，扩展类型走 project sheet）
    if dtype in IMPORTERS:
        list_url = url_for(f"ledger.{dtype}_list", project_id=project_id)
    else:
        sheet_view = REVERSE_SHEET_MAP.get(dtype)
        list_url = url_for(f"project.{sheet_view}") if sheet_view else url_for("project.dashboard")

    return render_template(
        "ledger/import.html",
        project=p,
        dtype=dtype,
        label=labels.get(dtype, dtype),
        list_url=list_url,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu or dtype,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "导入" + labels.get(dtype, dtype), "url": None},
        ),
        page_title="导入" + labels.get(dtype, dtype),
    )


def get_project(project_id):
    p = Project.query.get_or_404(project_id)
    if not current_user.is_admin:
        from app.models import UserProject
        if not UserProject.query.filter_by(user_id=current_user.id, project_id=project_id).first():
            from flask import abort
            abort(403)
    return p


def apply_date_filter(query, model, start, end, period=None, date_field="date"):
    date_col = getattr(model, date_field)
    if period:
        try:
            y, m = period.split('-')
            query = query.filter(
                extract('year', date_col) == int(y),
                extract('month', date_col) == int(m)
            )
        except Exception:
            pass
    if start:
        try:
            query = query.filter(date_col >= datetime.strptime(start, "%Y-%m-%d"))
        except Exception:
            pass
    if end:
        try:
            query = query.filter(date_col <= datetime.strptime(end, "%Y-%m-%d"))
        except Exception:
            pass
    return query


def _render_unit_values(items, unit, weight_field):
    """根据单位计算显示重量"""
    total = 0
    for i in items:
        raw = getattr(i, weight_field) or 0
        if unit == "kg":
            setattr(i, "display_weight", round(raw * 1000, 1))
            total += raw * 1000
        else:
            setattr(i, "display_weight", round(raw, 3))
            total += raw
    return round(total, 3 if unit == "T" else 1)


# ===== 进场台账 =====
@ledger_bp.route("/<int:project_id>/incoming")
@login_required
def incoming_list(project_id, active_menu="file_management", active_submenu="incoming"):
    p = get_project(project_id)
    unit = request.args.get("unit", "T")
    start, end, period = request.args.get("start"), request.args.get("end"), request.args.get("period")
    q = Incoming.query.filter_by(project_id=project_id)
    q = apply_date_filter(q, Incoming, start, end, period)
    pagination = paginate(q.order_by(Incoming.date.desc()))
    items = pagination.items
    total = _render_unit_values(items, unit, "weigh_weight")
    return render_template(
        "ledger/incoming.html",
        project=p,
        items=items,
        pagination=pagination,
        total=total,
        unit=unit,
        start=start,
        end=end,
        period=period,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "进场台账", "url": None},
        ),
        page_title="进场台账",
    )


# ===== 调拨台账 =====
@ledger_bp.route("/<int:project_id>/transfer")
@login_required
def transfer_list(project_id, active_menu="file_management", active_submenu="transfer"):
    p = get_project(project_id)
    unit = request.args.get("unit", "T")
    start, end, period = request.args.get("start"), request.args.get("end"), request.args.get("period")
    q = Transfer.query.filter_by(project_id=project_id)
    q = apply_date_filter(q, Transfer, start, end, period)
    pagination = paginate(q.order_by(Transfer.date.desc()))
    items = pagination.items
    total = _render_unit_values(items, unit, "weight")
    return render_template(
        "ledger/transfer.html",
        project=p,
        items=items,
        pagination=pagination,
        total=total,
        unit=unit,
        start=start,
        end=end,
        period=period,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "调拨台账", "url": None},
        ),
        page_title="调拨台账",
    )


# ===== 措施筋台账 =====
@ledger_bp.route("/<int:project_id>/measure")
@login_required
def measure_list(project_id, active_menu="file_management", active_submenu="measure"):
    p = get_project(project_id)
    unit = request.args.get("unit", "T")
    start, end, period = request.args.get("start"), request.args.get("end"), request.args.get("period")
    q = MeasureRebar.query.filter_by(project_id=project_id)
    q = apply_date_filter(q, MeasureRebar, start, end, period)
    pagination = paginate(q.order_by(MeasureRebar.date.desc()))
    items = pagination.items
    total = 0
    for i in items:
        raw = i.weight_kg or 0
        if unit == "T":
            setattr(i, "display_w", round(raw / 1000, 3))
            total += raw / 1000
        else:
            setattr(i, "display_w", round(raw, 1))
            total += raw
    total = round(total, 3 if unit == "T" else 1)
    return render_template(
        "ledger/measure.html",
        project=p,
        items=items,
        pagination=pagination,
        total=total,
        unit=unit,
        start=start,
        end=end,
        period=period,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "措施筋台账", "url": None},
        ),
        page_title="措施筋台账",
    )


# ===== 盘点表 =====
@ledger_bp.route("/<int:project_id>/inventory")
@login_required
def inventory_list(project_id, active_menu="file_management", active_submenu="inventory"):
    p = get_project(project_id)
    unit = request.args.get("unit", "T")
    start, end, period = request.args.get("start"), request.args.get("end"), request.args.get("period")
    q = Inventory.query.filter_by(project_id=project_id)
    q = apply_date_filter(q, Inventory, start, end, period, date_field="inventory_date")
    pagination = paginate(q.order_by(Inventory.inventory_date.desc()))
    items = pagination.items
    total = _render_unit_values(items, unit, "total_weight")
    return render_template(
        "ledger/inventory.html",
        project=p,
        items=items,
        pagination=pagination,
        total=total,
        unit=unit,
        start=start,
        end=end,
        period=period,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "盘点表", "url": None},
        ),
        page_title="盘点表",
    )


# ===== 废料台账 =====
@ledger_bp.route("/<int:project_id>/waste")
@login_required
def waste_list(project_id, active_menu="file_management", active_submenu="waste"):
    p = get_project(project_id)
    unit = request.args.get("unit", "T")
    start, end, period = request.args.get("start"), request.args.get("end"), request.args.get("period")
    q = Waste.query.filter_by(project_id=project_id)
    q = apply_date_filter(q, Waste, start, end, period, date_field="process_date")
    pagination = paginate(q.order_by(Waste.process_date.desc()))
    items = pagination.items
    total = _render_unit_values(items, unit, "waste_weight")
    return render_template(
        "ledger/waste.html",
        project=p,
        items=items,
        pagination=pagination,
        total=total,
        unit=unit,
        start=start,
        end=end,
        period=period,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "废料台账", "url": None},
        ),
        page_title="废料台账",
    )


# ===== 形象进度 =====
@ledger_bp.route("/<int:project_id>/progress")
@login_required
def progress_view(project_id, active_menu="file_management", active_submenu="progress"):
    p = get_project(project_id)
    # 获取形象进度汇总
    from app.models.building_progress import BuildingProgress
    buildings = db.session.query(
        BuildingProgress.building_name,
        func.sum(BuildingProgress.model_total).label("model_total"),
        func.sum(BuildingProgress.progress_qty).label("progress_qty"),
        func.sum(BuildingProgress.total_weight).label("total_weight"),
        func.max(BuildingProgress.progress_status).label("status"),
    ).filter_by(project_id=project_id).group_by(BuildingProgress.building_name).all()

    total_model = round(sum(b.model_total or 0 for b in buildings), 1)
    total_progress = round(sum(b.progress_qty or 0 for b in buildings), 1)

    return render_template(
        "ledger/progress.html",
        project=p,
        buildings=buildings,
        total_model=total_model,
        total_progress=total_progress,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "形象进度", "url": None},
        ),
        page_title="形象进度",
    )


@ledger_bp.route("/<int:project_id>/progress/<building_name>")
@login_required
def progress_detail(project_id, building_name, active_menu="file_management", active_submenu="progress"):
    p = get_project(project_id)
    from app.models.building_progress import BuildingProgress
    items = BuildingProgress.query.filter_by(
        project_id=project_id, building_name=building_name
    ).order_by(BuildingProgress.floor_name, BuildingProgress.component_type).all()
    total = round(sum(i.total_weight or 0 for i in items), 1)
    return render_template(
        "ledger/progress_detail.html",
        project=p,
        building=building_name,
        items=items,
        total=total,
        current_project=p,
        active_menu=active_menu,
        active_submenu=active_submenu,
        breadcrumbs=_bc(
            {"name": "项目管理", "url": url_for("admin.project_list")},
            {"name": p.name, "url": url_for("project.detail", project_id=p.id)},
            {"name": "钢筋台账", "url": None},
            {"name": "形象进度", "url": url_for("ledger.progress_view", project_id=p.id)},
            {"name": building_name, "url": None},
        ),
        page_title=building_name,
    )


# ===== 结余率分析导出 =====
@ledger_bp.route("/<int:project_id>/export-analysis")
@login_required
def export_analysis(project_id):
    """生成结余率分析Excel（按10号模板格式）"""
    import openpyxl as _xl
    from flask import send_file, current_app
    from openpyxl.styles import Font as _Font, Alignment as _Align, Border as _Border, Side as _Side, PatternFill as _Fill
    from app.services.analysis_service import recalc_project_analysis

    p = get_project(project_id)
    pa = recalc_project_analysis(project_id)
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "结余率分析"
    thin = _Side(style="thin")
    hf = _Font(bold=True, size=12)
    hfill = _Fill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    b = _Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{p.name}·钢筋精细化管理分析"
    ws["A1"].font = _Font(bold=True, size=14,
                          name="Microsoft YaHei")

    for j, h in enumerate(["指标", "项目", "数值(T)", "说明", "备注"], 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = _Font(bold=True, color="FFFFFF", name="Microsoft YaHei")
        c.fill = hfill
        c.border = b

    rows = [
        ("①", "对甲结算量", float(pa.contract_qty or 0), "合同约定结算量", ""),
        ("②", "进场量", float(pa.incoming_qty or 0), "钢筋进场总量", ""),
        ("③", "剩余量", float(pa.remaining_qty or 0), "盘点剩余量", ""),
        ("④", "废料量", float(pa.waste_qty or 0), "废料处理量", ""),
        ("⑤", "调拨量", float(pa.transfer_qty or 0), "调拨出量", ""),
        ("⑥", "非预算收入", float(pa.non_budget_use_qty or 0), "非预算收入使用量", ""),
        ("⑦", "使用量", float(pa.usage_qty or 0), "②-③-④-⑤-⑥", ""),
        ("⑧", "节约量", float(pa.saved_qty or 0), "①-⑦", ""),
        ("⑨", "结余率", f"{float(pa.balance_rate or 0)}%", "⑧/①×100%", "%"),
    ]
    for i, (idx, name, val, desc, unit) in enumerate(rows, 4):
        for j, v in enumerate([idx, name, val, desc, unit], 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = b
            if j == 3 and isinstance(v, (int, float)):
                c.number_format = "0.000"

    for col, w in [("A", 8), ("B", 18), ("C", 16), ("D", 30), ("E", 10)]:
        ws.column_dimensions[col].width = w

    ed = os.path.join(current_app.config["EXPORT_FOLDER"], "analysis")
    os.makedirs(ed, exist_ok=True)
    fn = f"{p.name}_结余率分析_{datetime.utcnow().strftime('%Y%m%d%H%M')}.xlsx"
    fp = os.path.join(ed, fn)
    wb.save(fp)
    return send_file(fp, as_attachment=True, download_name=fn)


# ===== 形象进度 CRUD =====
@ledger_bp.route("/<int:project_id>/progress/add", methods=["POST"])
@login_required
def progress_add(project_id):
    from app.models.building_progress import BuildingProgress
    db.session.add(BuildingProgress(
        project_id=project_id,
        building_name=request.form["building_name"],
        floor_name=request.form.get("floor_name", ""),
        component_type=request.form.get("component_type", ""),
        progress_status=request.form.get("status", "active"),
        model_total=float(request.form.get("model_total", 0) or 0),
        progress_qty=float(request.form.get("progress_qty", 0) or 0),
        total_weight=float(request.form.get("total_weight", 0) or 0),
        record_date=datetime.utcnow().date(),
    ))
    db.session.commit()
    flash("进度数据已添加。", "success")
    return redirect(url_for("ledger.progress_view", project_id=project_id))


@ledger_bp.route("/<int:project_id>/progress/<int:pid>/edit", methods=["POST"])
@login_required
def progress_edit(project_id, pid):
    from app.models.building_progress import BuildingProgress
    bp = BuildingProgress.query.get_or_404(pid)
    bp.building_name = request.form.get("building_name", bp.building_name)
    bp.progress_status = request.form.get("status", bp.progress_status)
    bp.model_total = float(request.form.get("model_total", bp.model_total or 0) or 0)
    bp.progress_qty = float(request.form.get("progress_qty", bp.progress_qty or 0) or 0)
    bp.total_weight = float(request.form.get("total_weight", bp.total_weight or 0) or 0)
    bp.record_date = datetime.utcnow().date()
    db.session.commit()
    flash("已更新。", "success")
    return redirect(url_for("ledger.progress_view", project_id=project_id))


@ledger_bp.route("/<int:project_id>/progress/<int:pid>/delete", methods=["POST"])
@login_required
def progress_delete(project_id, pid):
    from app.models.building_progress import BuildingProgress
    BuildingProgress.query.filter_by(id=pid).delete()
    db.session.commit()
    flash("已删除。", "success")
    return redirect(url_for("ledger.progress_view", project_id=project_id))
