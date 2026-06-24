# ============================================
# 配料单工程量导出服务
# ============================================
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app import db
from app.models.bom import Building, Floor, Area, Component, RebarDetail
from sqlalchemy import func

DIAMETER_COLS = [6, 8, 10, 12, 14, 16, 18, 20, 22, 25, 28, 32]


def export_project_summary(project_id: int, output_dir: str) -> str:
    """导出项目配料单汇总 Excel"""
    wb = openpyxl.Workbook()

    # ===== Sheet 1: 按构件汇总 =====
    ws1 = wb.active
    ws1.title = "构件汇总"

    thin = Side(style="thin")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # 表头
    headers = ["楼栋", "楼层", "区域", "构件名称", "构件类型", "施工状态"] + [f"{d}mm" for d in DIAMETER_COLS] + ["总重量(T)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # 数据
    buildings = Building.query.filter_by(project_id=project_id).order_by(Building.name).all()
    row = 2
    for b in buildings:
        for f in b.floors.order_by(Floor.sort_order, Floor.name).all():
            for a in f.areas.order_by(Area.name).all():
                for c in a.components.order_by(Component.name).all():
                    wbd = c.weight_by_diameter
                    ws1.cell(row=row, column=1, value=b.name).border = border
                    ws1.cell(row=row, column=2, value=f.name).border = border
                    ws1.cell(row=row, column=3, value=a.name).border = border
                    ws1.cell(row=row, column=4, value=c.name).border = border
                    ws1.cell(row=row, column=5, value=c.component_type).border = border
                    ws1.cell(row=row, column=6, value=c.status).border = border
                    for i, d in enumerate(DIAMETER_COLS):
                        v = wbd.get(d, 0)
                        if v > 0:
                            cell = ws1.cell(row=row, column=7 + i, value=round(v, 3))
                            cell.number_format = "0.000"
                            cell.border = border
                    total = c.total_weight
                    cell = ws1.cell(row=row, column=len(headers), value=round(total, 3))
                    cell.number_format = "0.000"
                    cell.font = Font(bold=True)
                    cell.border = border
                    row += 1

    # 列宽
    for i in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 10 if i > 6 else 15

    # ===== Sheet 2: 按楼层汇总 =====
    ws2 = wb.create_sheet("楼层汇总")
    floor_headers = ["楼层", "构件类型"] + [f"{d}mm" for d in DIAMETER_COLS] + ["总重量(T)"]
    for col, h in enumerate(floor_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 2
    for b in buildings:
        for f in b.floors.order_by(Floor.sort_order, Floor.name).all():
            # 汇总该楼层所有构件
            floor_data = {}
            for a in f.areas.all():
                for c in a.components.all():
                    ct = c.component_type
                    if ct not in floor_data:
                        floor_data[ct] = {}
                    for dia, w in c.weight_by_diameter.items():
                        floor_data[ct][dia] = floor_data[ct].get(dia, 0) + w

            for ct, dia_weights in floor_data.items():
                ws2.cell(row=row, column=1, value=f.name).border = border
                ws2.cell(row=row, column=2, value=ct).border = border
                ft = 0
                for i, d in enumerate(DIAMETER_COLS):
                    v = dia_weights.get(d, 0)
                    if v > 0:
                        cell = ws2.cell(row=row, column=3 + i, value=round(v, 3))
                        cell.number_format = "0.000"
                        cell.border = border
                    ft += v
                cell = ws2.cell(row=row, column=len(floor_headers), value=round(ft, 3))
                cell.number_format = "0.000"
                cell.font = Font(bold=True)
                cell.border = border
                row += 1

    for i in range(1, len(floor_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 10 if i > 2 else 15

    # 保存
    filename = f"配料单汇总_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
