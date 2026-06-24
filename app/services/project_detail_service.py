# ============================================
# 钢筋精细化管理平台 — 项目详情页数据聚合服务
# ============================================
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy import extract, func as sql_func

from app import db
from app.models import (
    Project,
    ProjectAnalysis,
    User,
    UserProject,
)
from app.models.bom import Area, Building, Component, Floor, RebarDetail
from app.models.business import ProjectAttachment, ProjectCost, ProjectProfitLoss
from app.models.rebar import (
    Incoming,
    Transfer,
    MeasureRebar,
    Inventory,
    Waste,
)
from app.models.building_progress import BuildingProgress


# ---------- 角色映射 ----------
MEMBER_ROLE_MAP = {
    "engineer": {"label": "精管工程师", "color": "primary"},
    "reviewer": {"label": "审核员", "color": "warning"},
    "viewer": {"label": "查看员", "color": "info"},
}


def _safe_float(value, precision=3):
    """安全转 float 并保留精度"""
    try:
        return round(float(value or 0), precision)
    except (TypeError, ValueError):
        return 0.0


def _safe_decimal(value, precision=2):
    """安全转 decimal/float"""
    try:
        return round(float(value or 0), precision)
    except (TypeError, ValueError):
        return 0.0


# ================= Tab1: 项目概况 =================

def get_project_overview(project_id):
    """项目基础信息 + 甲方/承接公司 + 附件 + ProjectAnalysis 指标"""
    project = Project.query.get_or_404(project_id)

    analysis = ProjectAnalysis.query.filter_by(project_id=project_id).first()
    if analysis:
        stats = {
            "incoming_total": _safe_float(analysis.incoming_qty),
            "usage_total": _safe_float(analysis.usage_qty),
            "balance_rate": _safe_float(analysis.balance_rate, 2),
            "loss_rate": _safe_float(
                (analysis.waste_qty / analysis.incoming_qty * 100)
                if analysis.incoming_qty else 0,
                2,
            ),
            "saved_total": _safe_float(analysis.saved_qty),
            "contract_total": _safe_float(analysis.contract_qty),
        }
    else:
        stats = {
            "incoming_total": 0.0,
            "usage_total": 0.0,
            "balance_rate": 0.0,
            "loss_rate": 0.0,
            "saved_total": 0.0,
            "contract_total": 0.0,
        }

    # 封面图
    cover = (
        ProjectAttachment.query.filter_by(project_id=project_id, is_cover=True)
        .order_by(ProjectAttachment.created_at.desc())
        .first()
    )
    # 最多 6 张效果图
    attachments = (
        ProjectAttachment.query.filter_by(project_id=project_id)
        .order_by(ProjectAttachment.is_cover.desc(), ProjectAttachment.created_at.desc())
        .limit(6)
        .all()
    )

    return {
        "project": project,
        "client_name": project.client_unit.name if project.client_unit else "-",
        "contractor_name": project.contractor.name if project.contractor else "-",
        "stats": stats,
        "cover": cover,
        "attachments": attachments,
    }


# ================= Tab2: 人员管理 =================

def get_project_members(project_id, role=None, search=None):
    """获取项目成员列表（含用户信息）"""
    query = (
        UserProject.query.filter_by(project_id=project_id)
        .join(User, UserProject.user_id == User.id)
    )

    if role:
        query = query.filter(UserProject.role == role)

    if search:
        search_like = f"%{search}%"
        query = query.filter(
            sql_func.or_(
                User.display_name.ilike(search_like),
                User.username.ilike(search_like),
            )
        )

    return query.order_by(UserProject.created_at.desc()).all()


def add_project_member(project_id, user_id, role="engineer"):
    """添加项目成员"""
    # 检查是否已存在
    exists = UserProject.query.filter_by(
        project_id=project_id, user_id=user_id
    ).first()
    if exists:
        raise ValueError("该用户已是项目成员")

    user = User.query.get(user_id)
    if not user:
        raise ValueError("用户不存在")

    up = UserProject(
        project_id=project_id,
        user_id=user_id,
        role=role,
    )
    db.session.add(up)
    db.session.commit()
    return up


def remove_project_member(project_id, user_id):
    """移除项目成员"""
    up = UserProject.query.filter_by(
        project_id=project_id, user_id=user_id
    ).first()
    if not up:
        raise ValueError("成员不存在")
    db.session.delete(up)
    db.session.commit()
    return up


def get_available_users(project_id):
    """获取可加入项目的用户（排除已在项目中的用户）"""
    member_ids = [
        up.user_id
        for up in UserProject.query.filter_by(project_id=project_id).all()
    ]
    return (
        User.query.filter(User.is_active == True)
        .filter(~User.id.in_(member_ids) if member_ids else True)
        .order_by(User.display_name)
        .all()
    )


# ================= Tab3: 数据台账 =================

def get_ledger_summary(project_id):
    """7 个台账的汇总统计"""
    # 进场台账
    incoming_count = Incoming.query.filter_by(project_id=project_id).count()
    incoming_weight = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(Incoming.weigh_weight), 0))
        .filter_by(project_id=project_id)
        .scalar()
    )

    # 调拨台账
    transfer_count = Transfer.query.filter_by(project_id=project_id).count()
    transfer_weight = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(Transfer.weight), 0))
        .filter_by(project_id=project_id)
        .scalar()
    )

    # 措施筋台账
    measure_count = MeasureRebar.query.filter_by(project_id=project_id).count()
    measure_weight_kg = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(MeasureRebar.weight_kg), 0))
        .filter_by(project_id=project_id)
        .scalar(),
        precision=1,
    )

    # 盘点台账
    inventory_count = Inventory.query.filter_by(project_id=project_id).count()
    inventory_weight = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(Inventory.total_weight), 0))
        .filter_by(project_id=project_id)
        .scalar()
    )

    # 废料台账
    waste_count = Waste.query.filter_by(project_id=project_id).count()
    waste_weight = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(Waste.waste_weight), 0))
        .filter_by(project_id=project_id)
        .scalar()
    )

    # 形象进度
    bp_count = BuildingProgress.query.filter_by(project_id=project_id).count()
    bp_agg = db.session.query(
        sql_func.coalesce(sql_func.sum(BuildingProgress.model_total), 0),
        sql_func.coalesce(sql_func.sum(BuildingProgress.progress_qty), 0),
    ).filter_by(project_id=project_id).first()
    model_total = _safe_float(bp_agg[0], 3)
    progress_qty = _safe_float(bp_agg[1], 3)
    progress_rate = round((progress_qty / model_total * 100), 1) if model_total > 0 else 0.0

    # 配料单：楼栋数 / 构件数 / 已浇筑数
    building_count = Building.query.filter_by(project_id=project_id).count()
    component_count = (
        Component.query.join(Area).join(Floor).join(Building)
        .filter(Building.project_id == project_id)
        .count()
    )
    poured_count = (
        Component.query.join(Area).join(Floor).join(Building)
        .filter(Building.project_id == project_id, Component.status == "poured")
        .count()
    )
    bom_tonnage = _safe_float(
        db.session.query(sql_func.coalesce(sql_func.sum(RebarDetail.weight), 0))
        .join(Component).join(Area).join(Floor).join(Building)
        .filter(Building.project_id == project_id)
        .scalar()
    )

    return {
        "incoming": {"count": incoming_count, "weight": incoming_weight},
        "transfer": {"count": transfer_count, "weight": transfer_weight},
        "measure": {"count": measure_count, "weight_kg": measure_weight_kg},
        "inventory": {"count": inventory_count, "weight": inventory_weight},
        "waste": {"count": waste_count, "weight": waste_weight},
        "progress": {
            "count": bp_count,
            "building_count": building_count,
            "progress_rate": progress_rate,
        },
        "bom": {
            "count": component_count,
            "poured_count": poured_count,
            "building_count": building_count,
            "tonnage": bom_tonnage,
        },
    }


# ================= Tab4: 结算分析 =================

def get_settlement_analysis(project_id):
    """ProjectAnalysis 结算数据"""
    analysis = ProjectAnalysis.query.filter_by(project_id=project_id).first()
    if not analysis:
        analysis = ProjectAnalysis(project_id=project_id)

    incoming = _safe_float(analysis.incoming_qty)
    contract = _safe_float(analysis.contract_qty)
    usage = _safe_float(analysis.usage_qty)
    saved = _safe_float(analysis.saved_qty)
    waste = _safe_float(analysis.waste_qty)

    loss_rate = round((waste / incoming * 100), 2) if incoming > 0 else 0.0
    efficiency_rate = round((saved / contract * 100), 2) if contract > 0 else 0.0

    return {
        "contract_qty": contract,
        "incoming_qty": incoming,
        "usage_qty": usage,
        "saved_qty": saved,
        "balance_rate": _safe_float(analysis.balance_rate, 2),
        "loss_rate": loss_rate,
        "efficiency_rate": efficiency_rate,
        # 明细
        "model_qty": _safe_float(analysis.struct_qty),
        "measure_qty": _safe_float(analysis.measure_qty),
        "inventory_qty": _safe_float(analysis.remaining_qty),
        "waste_qty": waste,
        "transfer_qty": _safe_float(analysis.transfer_qty),
    }


# ================= Tab5: 盈亏分析 =================

def get_profit_analysis(project_id):
    """ProjectProfitLoss + ProjectCost 数据"""
    pl = ProjectProfitLoss.query.filter_by(project_id=project_id).first()
    if not pl:
        pl = ProjectProfitLoss(project_id=project_id)

    costs = (
        ProjectCost.query.filter_by(project_id=project_id)
        .order_by(ProjectCost.cost_date.desc())
        .all()
    )

    # 按类别汇总成本
    category_summary = {}
    total_cost = 0.0
    for cost in costs:
        cat = cost.cost_category
        amount = _safe_decimal(cost.amount, 2)
        category_summary[cat] = category_summary.get(cat, 0.0) + amount
        total_cost += amount

    # 成本类别中文映射
    category_labels = {
        "labor": "人工费",
        "material": "材料费",
        "equipment": "机械费",
        "transport": "运输费",
        "management": "管理费",
        "other": "其他",
    }
    cost_breakdown = [
        {
            "category": category_labels.get(cat, cat),
            "amount": round(amount, 2),
            "ratio": round(amount / total_cost * 100, 1) if total_cost > 0 else 0.0,
        }
        for cat, amount in category_summary.items()
    ]

    rebar_income = _safe_decimal(pl.rebar_income, 2)
    other_income = _safe_decimal(pl.other_income, 2)
    total_income = _safe_decimal(pl.total_income, 2) or (rebar_income + other_income)
    net_profit = _safe_decimal(pl.net_profit, 2)
    profit_rate = _safe_decimal(pl.profit_rate, 2)

    # 月度成本趋势（最近12个月）
    monthly_costs = []
    now = datetime.utcnow()
    for i in range(11, -1, -1):
        d = now.replace(day=1) - timedelta(days=i * 30)
        # 简化：按月份近似，取该月所在年月
        y, m = d.year, d.month
        amount = _safe_decimal(
            db.session.query(sql_func.coalesce(sql_func.sum(ProjectCost.amount), 0))
            .filter(
                ProjectCost.project_id == project_id,
                extract("year", ProjectCost.cost_date) == y,
                extract("month", ProjectCost.cost_date) == m,
            )
            .scalar(),
            2,
        )
        monthly_costs.append({"label": f"{m}月", "amount": amount})

    return {
        "rebar_income": rebar_income,
        "other_income": other_income,
        "total_income": total_income,
        "total_cost": total_cost,
        "net_profit": net_profit,
        "profit_rate": profit_rate,
        "cost_breakdown": cost_breakdown,
        "monthly_costs": monthly_costs,
        "rebar_unit_price": _safe_decimal(pl.rebar_unit_price, 2),
    }


def export_settlement_analysis(project_id: int, output_dir: str) -> str:
    """导出项目结算分析 Excel"""
    project = Project.query.get_or_404(project_id)
    data = get_settlement_analysis(project_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "结算分析"

    thin = Side(style="thin")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = ["指标", "数值", "单位"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    rows = [
        ["项目名称", project.name, ""],
        ["施工图预算量", data["contract_qty"], "T"],
        ["进场总量", data["incoming_qty"], "T"],
        ["消耗总量", data["usage_qty"], "T"],
        ["节约量", data["saved_qty"], "T"],
        ["结余率", data["balance_rate"], "%"],
        ["损耗率", data["loss_rate"], "%"],
        ["效益率", data["efficiency_rate"], "%"],
        ["模型量", data["model_qty"], "T"],
        ["措施量", data["measure_qty"], "T"],
        ["剩余量", data["inventory_qty"], "T"],
        ["废料量", data["waste_qty"], "T"],
        ["调拨量", data["transfer_qty"], "T"],
    ]

    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if c_idx == 2 and isinstance(value, float):
                cell.number_format = "0.000"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10

    filename = f"settlement_{project_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
